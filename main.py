import time

from Bio import SeqIO
from mongoengine import connect
from db.ExecutionTime import ExecutionTime
from db.SequenceDocument import SequenceDocument
from db.SequenceStats import count_wild_type, count_distance_distribution, sequence_stats, get_distribution_stats, \
    count_paired_distance_distribution, count_nucleoid_stats, init_seq_position_stats, reset_seq_position_stats, \
    count_distance
from excel.res_xls import create_res_xls
from openpyxl import Workbook
from dotenv import load_dotenv
import os

load_dotenv()
connect(host=os.environ.get("MONGO_HOST"))

regionMap = {
    "Ivano-Frankovskaya": "IF",
    "Belgorodskaya oblast', Krasnogvardeysky rayon": "BK",
    "Belgorodskaya oblast', Grayvoronsky rayon": "BG",
    "L'vovskaya oblast', Stryi": "ST",
    "Cherkasskaya": "CH",
    "Khmelnitskaya": "KHM",
}

otherRegionMap = {
    "Donetskaya": "DO",
    "Odesskaya": "OD",
    "Rovenskaya": "RO",
    "Kharkovskaya": "KHA",
    "Zhitomirskaya": "ZH",
    "Sumskaya": "SU",
    "Belgorodskaya oblast' unspecified": "B",
    "undefined: UND": "UND"
}


# def countATGC(db):
#     Ac = db.count("A")
#     Tc = db.count("T")
#     Gc = db.count("G")
#     Cc = db.count("C")
#     Lc = (Ac + Tc + Gc + Cc)
#     GCc = ((Gc + Cc) / (Ac + Tc + Gc + Cc + 0.0)) * 100
#     ATc = ((Ac + Tc) / (Ac + Tc + Gc + Cc + 0.0)) * 100
#     return Ac, Tc, Gc, Cc, Lc, GCc, ATc

def reset_sequence_documents_regions():
    for seq_document in SequenceDocument.objects(length=377):
        seq_document.delete()


def parce_base_sequence(filePath, file_format):
    for record in SeqIO.parse(filePath, file_format):
        seqId = str(record.id)
        seq = str(record.seq.upper())
        document = SequenceDocument(
            version=seqId,
            length=len(seq),
            fasta=record.format('fasta'),
            sequence=seq,
            name=record.name
        )
        document.save()
        return document.id


def query_base_sequence(name):
    return SequenceDocument.objects(name=name)[0]


def query_not_base_sequences():
    length = 16569
    cursor = SequenceDocument.objects(length__ne=length)
    records = []
    for r in cursor:
        records.append(r)
    del cursor
    return records


def load_all_sequences(filename, file_format):
    start_time = time.time()
    for seqRecord in SeqIO.parse(filename, file_format):

        seq = seqRecord.seq.upper()
        location = []
        region = ''
        for feature in seqRecord.features:
            if feature.type == 'source':
                start = feature.location.nofuzzy_start
                end = feature.location.nofuzzy_end
                country = feature.qualifiers.get('country', ['undefined: UND'])
                location = [start, end]
                region = country[0]
        # count nucleoid stats if needed
        # reset stats first
        # count_nucleoid_stats(seq)
        document = SequenceDocument(
            version=str(seqRecord.id),
            length=len(seq),
            fasta=seqRecord.format('fasta'),
            sequence=str(seq),
            location=location,
            region=region
        )
        document.save()

    load_time = ExecutionTime(name="Sequence Load Time", seconds=time.time() - start_time)
    load_time.save()


def get_sequences_by_region(region):
    return SequenceDocument.objects(region__contains=region).only('sequence')


def save_stats_db(sequence):
    rcrs_base_sequence = query_base_sequence('NC_012920')
    rsrs_base_sequence = query_base_sequence('RSRS')
    doc = SequenceDocument.objects(version=sequence.version)
    rcrs_distance = count_distance(sequence.sequence, rcrs_base_sequence.sequence[16023:16400])
    doc.update(distance_to_rCRS=rcrs_distance)
    rsrs_distance = count_distance(sequence.sequence, rsrs_base_sequence.sequence[16023:16400])
    doc.update(distance_to_RSRS=rsrs_distance)


def query_normal_length():
    length = 377
    cursor = SequenceDocument.objects(length=length)
    records = []
    for r in cursor:
        records.append(r)
    del cursor
    return records

def create_sheet_with_results(workbook, region, sequences):
    # query rCRS base db
    rcrs_base_sequence = query_base_sequence('NC_012920')
    # parse RSRS base db
    rsrs_base_sequence = query_base_sequence('RSRS')
    wild_type = count_wild_type()
    wild_type_rcrs_distance = count_distance_distribution([wild_type], rcrs_base_sequence.sequence[16023:16400])[0]
    wild_type_rsrs_distance = count_distance_distribution([wild_type], rsrs_base_sequence.sequence[16023:16400])[0]

    rcrs_distance = count_distance_distribution(sequences, rcrs_base_sequence.sequence[16023:16400])
    rcrs_distance_stats = sequence_stats(rcrs_distance)
    rcrs_histogram, rcrs_histogram_parts = get_distribution_stats(rcrs_distance)

    rsrs_distance = count_distance_distribution(sequences, rsrs_base_sequence.sequence[16023:16400])
    rsrs_distance_stats = sequence_stats(rsrs_distance)
    rsrs_histogram, rsrs_histogram_parts = get_distribution_stats(rsrs_distance)

    wild_type_distance = count_distance_distribution(sequences, wild_type)
    wild_type_distance_stats = sequence_stats(wild_type_distance)
    wild_type_histogram, wild_type_histogram_parts = get_distribution_stats(wild_type_distance)

    paired_distance = count_paired_distance_distribution(sequences)
    paired_distance_stats = sequence_stats(paired_distance)
    paired_histogram, paired_histogram_parts = get_distribution_stats(paired_distance)

    other_stats = [str(wild_type), wild_type_rcrs_distance, wild_type_rsrs_distance, sum(rcrs_distance),
                   sum(rsrs_distance)]
    print(f'rCRS DISTANCE = {rcrs_distance}')
    print(f'rCRS DISTANCE histogram = {rcrs_histogram}')
    print(f'rCRS DISTANCE histogram parted = {rcrs_histogram_parts}')
    print(f'rCRS DISTANCE stats = {rcrs_distance_stats}')

    print(f'RSRS DISTANCE = {rsrs_distance}')
    print(f'RSRS DISTANCE stats = {rsrs_distance_stats}')

    print(f'wild type DISTANCE = {wild_type_distance}')
    print(f'wild type DISTANCE stats = {wild_type_distance_stats}')
    print("======================")

    regionMapped = regionMap.get(region, 'Other Regions')

    create_res_xls(
        workbook=workbook,
        rcrs_distribution=rcrs_histogram,
        rcrs_distribution_percent=rcrs_histogram_parts,
        rcrs_stats=rcrs_distance_stats,
        rsrs_distribution=rsrs_histogram,
        rsrs_distribution_percent=rsrs_histogram_parts,
        rsrs_stats=rsrs_distance_stats,
        wild_type_distribution=wild_type_histogram,
        wild_type_distribution_percent=wild_type_histogram_parts,
        wild_type_stats=wild_type_distance_stats,
        pair_distribution=paired_histogram,
        pair_distribution_percent=paired_histogram_parts,
        pair_stats=paired_distance_stats,
        other_stats=other_stats,
        region=regionMapped
    )


if __name__ == '__main__':
    # parse rCRS base db
    # parce_base_sequence("data/rCRS.gb", "genbank")

    # rsrs_id = parce_base_sequence("data/RSRS.fasta", "fasta")

    # Initialize position stats documents
    # init_seq_position_stats(377)
    # Load the db
    # reset_sequence_documents_regions()
    # load_all_sequences("data/sequence.gb", "genbank")
    # reset_seq_position_stats()

    # start_time = time.time()
    # nonBase = query_not_base_sequences()
    # for record in nonBase:
    #     count_nucleoid_stats(record)
    # load_time = ExecutionTime(name="Sequence Position Stats", seconds=time.time() - start_time)
    # load_time.save()
    #
    # queryNormalLenght = query_normal_length()
    # #
    # start_time = time.time()
    # for record in queryNormalLenght:
    #     save_stats_db(record)
    # load_time = ExecutionTime(name="Count Distances to Base", seconds=time.time() - start_time)
    # load_time.save()
    # start_time = time.time()
    workbook = Workbook()
    for region in regionMap.keys():
        sequences_obj = get_sequences_by_region(region)
        sequences = [sequence.sequence for sequence in sequences_obj]
        create_sheet_with_results(workbook, region, sequences)

    sequences = []
    for region in otherRegionMap.keys():
        sequences_obj = get_sequences_by_region(region)
        sequences.extend([sequence.sequence for sequence in sequences_obj])
    create_sheet_with_results(workbook, 'Other Regions', sequences)
    # load_time = ExecutionTime(name="Excel Creation", seconds=time.time() - start_time)
    # load_time.save()
