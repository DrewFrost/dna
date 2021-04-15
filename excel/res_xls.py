from openpyxl import Workbook, load_workbook

distance = 40


def create_res_xls(
    workbook,
    rcrs_distribution,
    rcrs_distribution_percent,
    rcrs_stats,
    rsrs_distribution,
    rsrs_distribution_percent,
    rsrs_stats,
    wild_type_distribution,
    wild_type_distribution_percent,
    wild_type_stats,
    pair_distribution,
    pair_distribution_percent,
    pair_stats,
    other_stats,
    region
):
    statistics_template = ['Мат сподів.', 'Сер. квадр. відхил.', 'Мода', 'Мін.', 'Макс.', 'Коеф. варіац.']

    sheet = workbook.create_sheet(region)

    sheet.cell(row=1, column=1, value='Відстань')
    for i in range(distance):
        sheet.cell(row=1, column=i + 2, value=i)

    # rCRS

    sheet.cell(row=2, column=1, value='Розподіл відносно базової rCRS')
    for i in range(len(rcrs_distribution)):
        sheet.cell(row=2, column=i + 2, value=rcrs_distribution[i])

    sheet.cell(row=3, column=1, value='Розподіл відносно базової rCRS (частка)')
    for i in range(len(rcrs_distribution_percent)):
        sheet.cell(row=3, column=i + 2, value=rcrs_distribution_percent[i])

    for i in range(len(statistics_template)):
        sheet.cell(row=4, column=i + 2, value=statistics_template[i])

    for i in range(len(rcrs_stats)):
        sheet.cell(row=5, column=i + 2, value=rcrs_stats[i])

    # RSRS

    sheet.cell(row=6, column=1, value='Розподіл відносно базової RSRS')
    for i in range(len(rsrs_distribution)):
        sheet.cell(row=6, column=i + 2, value=rsrs_distribution[i])

    sheet.cell(row=7, column=1, value='Розподіл відносно базової RSRS (частка)')
    for i in range(len(rsrs_distribution_percent)):
        sheet.cell(row=7, column=i + 2, value=rsrs_distribution_percent[i])

    for i in range(len(statistics_template)):
        sheet.cell(row=8, column=i + 2, value=statistics_template[i])

    for i in range(len(rsrs_stats)):
        sheet.cell(row=9, column=i + 2, value=rsrs_stats[i])

    # Wild type

    sheet.cell(row=10, column=1, value='Розподіл відносно дикого типу')
    for i in range(len(wild_type_distribution)):
        sheet.cell(row=10, column=i + 2, value=wild_type_distribution[i])

    sheet.cell(row=11, column=1, value='Розподіл відносно дикого типу (частка)')
    for i in range(len(wild_type_distribution_percent)):
        sheet.cell(row=11, column=i + 2, value=wild_type_distribution_percent[i])

    for i in range(len(statistics_template)):
        sheet.cell(row=12, column=i + 2, value=statistics_template[i])

    for i in range(len(wild_type_stats)):
        sheet.cell(row=13, column=i + 2, value=wild_type_stats[i])

    # Paired

    sheet.cell(row=14, column=1, value='Розподіл відносно попарних')
    for i in range(len(pair_distribution)):
        sheet.cell(row=14, column=i + 2, value=pair_distribution[i])

    sheet.cell(row=15, column=1, value='Розподіл відносно попарних (частка)')
    for i in range(len(pair_distribution_percent)):
        sheet.cell(row=15, column=i + 2, value=pair_distribution_percent[i])

    for i in range(len(statistics_template)):
        sheet.cell(row=16, column=i + 2, value=statistics_template[i])

    for i in range(len(pair_stats)):
        sheet.cell(row=17, column=i + 2, value=pair_stats[i])

    sheet.cell(row=18, column=2, value='Рядочок дикого типу')
    sheet.cell(row=19, column=2, value='Кількість поліморфізмів у дикого типу відносно базової rCRS')
    sheet.cell(row=20, column=2, value='Кількість поліморфізмів у дикого типу відносно базової RSRS')
    sheet.cell(row=22, column=2, value='Кількість поліморфізмів у популяції відносно базової rCRS')
    sheet.cell(row=23, column=2, value='Кількість поліморфізмів у популяції відносно базової RSRS')

    for i in range(0, 3):
        sheet.cell(row=i+18, column=3, value=other_stats[i])

    for i in range(3, 5):
        sheet.cell(row=i+19, column=3, value=other_stats[i])

    workbook.save(filename=f"../results.xlsx")